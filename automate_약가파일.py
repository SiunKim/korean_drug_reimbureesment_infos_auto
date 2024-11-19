import os
import pickle
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict

def generate_drug_info_sheet(compound_ids, variables):
    """Generate drug info sheet with enhanced styling and company count"""
    # List to store result data
    result_data = []
    # Add column names
    columns = ['투여', '분류', '주성분코드', '제품코드/주성분이름',
               '제품명', '업체명', '규격', '단위', '전일', '비고 - 급여 날짜', '업체 수']
    result_data.append(dict(zip(columns, columns)))
    # Process each compound id
    for compound_id in compound_ids:
        # Get compound information
        compound_info = variables['dict_main_compound_infos_total_repr'].get(compound_id, {})
        if len(compound_info) == 0:
            print(f"compound_id {compound_id} is not found in our database!")
            continue
        administration = compound_info.get('투여', '')
        compound_class = compound_info.get('분류', '')
        compound_name = compound_info.get('주성분이름', '')
        # Find products with this compound
        products_with_compound = []
        for product_id, compound_dates in \
            variables['compound_id_and_dates_by_product_id_all_period'].items():
            if any(compound_id == cd[0] for cd in compound_dates):
                products_with_compound.append(product_id)
        # Count unique companies
        companies = set()
        for product_id in products_with_compound:
            product_info = variables['product_info_by_id'].get(product_id, [{}])[0]
            companies.add(product_info[1])  # company name is at index 1
        company_count = len(companies)
        # Add compound row
        result_data.append({
            '투여': administration,
            '분류': compound_class,
            '주성분코드': compound_id,
            '제품코드/주성분이름': compound_name,
            '제품명': '',
            '업체명': '',
            '규격': '',
            '단위': '',
            '전일': '',
            '비고 - 급여 날짜': '',
            '업체 수': company_count
        })

        # Add product information
        for product_id in products_with_compound:
            product_info = variables['product_info_by_id'].get(product_id, [{}])[0]
            # Use first info
            product_name, company_name, size, unit, general_expert = product_info
            result_data.append({
                '투여': administration,
                '분류': compound_class,
                '주성분코드': compound_id,
                '제품코드/주성분이름': product_id,
                '제품명': product_name,
                '업체명': company_name,
                '규격': size,
                '단위': unit,
                '전일': general_expert,
                '비고 - 급여 날짜': variables['continuous_events_str_by_product_id'].get(product_id, ''),
                '업체 수': ''
            })
    # Create DataFrame
    df = pd.DataFrame(result_data)

    # Create Excel file and apply styling
    wb = Workbook()
    ws = wb.active
    # Write DataFrame to worksheet
    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Set column widths and alignments
    column_widths = [48, 48, 132, 190, 438, 113, 48, 48, 48, 463, 48]
    alignments = ['center', 'center', 'center', 'center',
                  'left', 'left', 'left', 'left', 'left', 'left', 'center']
    for i, (width, align) in enumerate(zip(column_widths, alignments), 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width / 7  # Convert to Excel width units
        for cell in ws[column_letter]:
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.font = Font(name='굴림')  # Set default font to 굴림
    # Apply styles to compound info rows
    light_green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Exclude header row
        if row[4].value == '':  # Product name is empty (compound info row)
            for cell in row:
                cell.fill = light_green_fill
                cell.font = Font(name='굴림', size=12)  # Increase font size by 2 points

            # Change style for '제품코드/주성분이름' cell
            row[3].alignment = Alignment(horizontal='left', vertical='center')
            row[3].font = Font(name='Verdana', size=12, bold=True)

    # Apply header row style
    for cell in ws[1]:
        cell.font = Font(name='굴림', bold=True)

    return wb

if __name__ == "__main__":
    # Load pickle data
    save_dir = "pickle_save_tmp_0911/variables_merged_from_2009_to_2024.pkl"
    with open(save_dir, 'rb') as f:
        variables = pickle.load(f)

    # Define compound IDs
    compound_ids = [
        '148630ALQ', '148631ALQ', '148601ATB', '148601ATD', '148602APD', 
        '148602ATB', '148602ATD', '148603ATB', '148604ATB', '190031ALQ', 
        '190001ATB', '190003ATD', '190004ATB', '190004ATD', '190005ATB', 
        '190006ATD', '224501ACH', '224503ACH', '224504ACH', '224505ACH', 
        '385203ACR', '385203ATR', '385204ACR', '385204ATR', '385205ACR', 
        '385205ATR', '165301ATB', '165303ATB', '165305ATB', '165310ATB', 
        '224506CPC', '224507CPC', '224508CPC', '643403CPC', '643404CPC'
    ]

    # Generate and save the Excel file
    wb = generate_drug_info_sheet(compound_ids, variables)
    wb.save('drug_info_sheet_styled_v3_치매_자동추출.xlsx')
    print("Excel file has been generated successfully.")
